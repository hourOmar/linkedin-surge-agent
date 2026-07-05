"""Loads the LinkedIn content calendar from an Excel workbook.

The calendar is organized as a sequence of week sections. Each week contains
up to four post slots ("Post 1".."Post 4"), and each slot may have an
embedded image, a caption/content cell, and a status cell nearby. Rather than
assuming fixed cell coordinates, this module locates "Week N" / "Post N"
labels wherever they appear in the sheet, which keeps it resilient to minor
layout changes (extra spacer rows, reordered columns, merged header cells).

Public API:
    load_content_calendar(path) -> CalendarData
    get_week(calendar, week)
    get_post(calendar, week, post)
    iter_posts(calendar)
"""

import re
from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

CalendarData = dict[str, dict[str, dict[str, Any]]]

CONTENT_CALENDAR_FILENAME = "linkedin_content_calendar.xlsx"

_WEEK_PATTERN = re.compile(r"^\s*week\D*(\d+)\s*$", re.IGNORECASE)
_POST_PATTERN = re.compile(r"^\s*post\D*(\d+)\s*$", re.IGNORECASE)

_HEADER_SEARCH_ROWS = 10
_CONTENT_HEADER_ALIASES = {"content", "caption", "copy", "post copy", "post content", "text"}
_STATUS_HEADER_ALIASES = {"status", "post status"}


class ContentCalendarLoadError(Exception):
    """Raised when the Excel content calendar cannot be read or parsed."""


def load_content_calendar(path: Path | str) -> CalendarData:
    """Load the content calendar workbook at `path` into a nested dict.

    Returns an empty dict if the file does not exist (the calendar is
    optional). Raises `ContentCalendarLoadError` if the file exists but
    cannot be parsed as a valid workbook.

    Shape of the result::

        {
            "Week 1": {
                "Post 1": {"content": "...", "status": "...", "image": {...} | None},
                "Post 2": {...},
            },
            "Week 2": {...},
        }
    """
    path = Path(path)

    if not path.is_file():
        return {}

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise ContentCalendarLoadError(f"Invalid Excel file {path.name}: {exc}") from exc

    sheet = workbook.active
    return _parse_sheet(sheet)


def get_week(calendar: CalendarData, week: str | int) -> dict[str, dict[str, Any]]:
    """Return all posts scheduled for a given week, e.g. get_week(kb, 1) or get_week(kb, "Week 1")."""
    return calendar.get(_normalize_label("week", week), {})


def get_post(calendar: CalendarData, week: str | int, post: str | int) -> dict[str, Any] | None:
    """Return a single post, e.g. get_post(kb, 1, 2) for Week 1 / Post 2."""
    return get_week(calendar, week).get(_normalize_label("post", post))


def iter_posts(calendar: CalendarData) -> Iterator[tuple[str, str, dict[str, Any]]]:
    """Iterate over every planned post as (week_label, post_label, post_data) tuples."""
    for week_label, posts in calendar.items():
        for post_label, post_data in posts.items():
            yield week_label, post_label, post_data


def _normalize_label(kind: str, value: str | int) -> str:
    if isinstance(value, int):
        number = value
    else:
        match = re.search(r"\d+", str(value))
        number = int(match.group()) if match else value
    return f"{kind.capitalize()} {number}"


def _parse_sheet(sheet: Worksheet) -> CalendarData:
    merged_values = _resolve_merged_cells(sheet)
    content_col, status_col = _find_header_columns(sheet)
    images_by_row = _index_images_by_row(sheet)

    calendar: CalendarData = {}
    current_week: str | None = None

    for row in sheet.iter_rows():
        row_values = [merged_values.get((cell.row, cell.column), cell.value) for cell in row]

        week_label = _match_label(row_values, _WEEK_PATTERN, "week")
        if week_label is not None:
            current_week = week_label
            calendar.setdefault(current_week, {})

        post_label = _match_label(row_values, _POST_PATTERN, "post")
        if post_label is None or current_week is None:
            continue

        row_num = row[0].row
        calendar[current_week][post_label] = {
            "content": _cell_text(row_values, content_col),
            "status": _cell_text(row_values, status_col),
            "image": images_by_row.get(row_num),
        }

    return calendar


def _resolve_merged_cells(sheet: Worksheet) -> dict[tuple[int, int], Any]:
    """Map every (row, col) inside a merged range to that range's top-left value."""
    resolved: dict[tuple[int, int], Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        top_left_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                resolved[(row, col)] = top_left_value
    return resolved


def _find_header_columns(sheet: Worksheet) -> tuple[int | None, int | None]:
    """Search the first few rows for columns labeled Content/Caption and Status."""
    content_col: int | None = None
    status_col: int | None = None

    for row in sheet.iter_rows(max_row=_HEADER_SEARCH_ROWS):
        for cell in row:
            label = str(cell.value).strip().lower() if cell.value is not None else ""
            if content_col is None and label in _CONTENT_HEADER_ALIASES:
                content_col = cell.column
            if status_col is None and label in _STATUS_HEADER_ALIASES:
                status_col = cell.column

    return content_col, status_col


def _index_images_by_row(sheet: Worksheet) -> dict[int, dict[str, Any]]:
    """Map each 1-indexed row number to the image anchored there, if any."""
    images_by_row: dict[int, dict[str, Any]] = {}
    for image in getattr(sheet, "_images", []):
        row = image.anchor._from.row + 1
        images_by_row[row] = {"format": image.format, "data": image._data()}
    return images_by_row


def _match_label(row_values: list[Any], pattern: re.Pattern[str], kind: str) -> str | None:
    for value in row_values:
        if value is None:
            continue
        match = pattern.match(str(value))
        if match:
            return f"{kind.capitalize()} {int(match.group(1))}"
    return None


def _cell_text(row_values: list[Any], column: int | None) -> str | None:
    if column is None or column > len(row_values):
        return None
    value = row_values[column - 1]
    return str(value).strip() if value is not None else None
