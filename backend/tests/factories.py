"""Shared test data builders (not collected by pytest as a test module)."""

import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


def _png_bytes() -> io.BytesIO:
    buf = io.BytesIO()
    PILImage.new("RGB", (4, 4), color=(255, 0, 0)).save(buf, format="PNG")
    buf.seek(0)
    return buf


def build_sample_calendar(path):
    """Build a workbook mirroring the real layout: a merged 'Week N' cell spanning
    several rows, one row per post, with content/status columns and an image
    anchored next to the first post of week 1. Week 1 only has 3 of its 4 post
    slots filled, matching how the real calendar leaves some slots blank."""
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Week"
    ws["B1"] = "Post"
    ws["C1"] = "Content"
    ws["D1"] = "Status"

    ws["A2"] = "Week 1"
    ws.merge_cells("A2:A5")
    ws["B2"] = "Post 1"
    ws["C2"] = "Caption for week 1 post 1"
    ws["D2"] = "Draft"
    ws.add_image(XLImage(_png_bytes()), "E2")

    ws["B3"] = "Post 2"
    ws["C3"] = "Caption for week 1 post 2"
    ws["D3"] = "Scheduled"

    ws["B4"] = "Post 3"
    ws["C4"] = "Caption for week 1 post 3"
    ws["D4"] = "Published"
    # Row 5 (Post 4) intentionally left blank.

    ws["A6"] = "Week 2"
    ws.merge_cells("A6:A7")
    ws["B6"] = "Post 1"
    ws["C6"] = "Caption for week 2 post 1"
    ws["D6"] = "Draft"

    ws["B7"] = "Post 2"
    ws["C7"] = "Caption for week 2 post 2"
    ws["D7"] = "Approved"

    wb.save(path)
    return path
